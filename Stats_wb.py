import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

wb = Workbook()
# ws = wb.active
ws1 = wb.create_sheet('Item Stats', 0)
ws2 = wb.create_sheet('969 Stats', 1)
ws3 = wb.create_sheet('CallNos', 2)
ws4 = wb.create_sheet('Print', 3)
ws5 = wb.create_sheet('Electronic', 4)

items = pd.read_csv('stats_out.csv', header=None)  #  the actual item-count statistics for each cataloger
items = dataframe_to_rows(items)
for r_idx, row in enumerate(items, 1):
    for c_idx, value in enumerate(row, 1):
         ws1.cell(row=r_idx, column=c_idx, value=value)
eBibedits = pd.read_csv('969_out.csv', header=None)  #  to count work done to bibs for electronic, etc, not counted by item edits
eBibedits = dataframe_to_rows(eBibedits)
for r_idx, row in enumerate(eBibedits, 1):
    for c_idx, value in enumerate(row, 1):
         ws2.cell(row=r_idx, column=c_idx, value=value)
callNos = pd.read_csv('CallNos_out.csv', header=None)  #  to list the accumulation of physical materials in call# ranges
callNos.sort_values(by=[0], axis=0, ascending=True, inplace=True)
callNos = callNos.reset_index(drop=True)
callNos = dataframe_to_rows(callNos)
for r_idx, row in enumerate(callNos, 1):
    for c_idx, value in enumerate(row, 1):
         ws3.cell(row=r_idx, column=c_idx, value=value)
gen_print = pd.read_csv('print_out.csv', header=None)  #  General Print Statistics original/copy
gen_print = dataframe_to_rows(gen_print)
for r_idx, row in enumerate(gen_print, 1):
    for c_idx, value in enumerate(row, 1):
         ws4.cell(row=r_idx, column=c_idx, value=value)
gen_elec = pd.read_csv('elec_out.csv', header=None)  #  General Print Statistics original/copy
gen_elec = dataframe_to_rows(gen_elec)
for r_idx, row in enumerate(gen_elec, 1):
    for c_idx, value in enumerate(row, 1):
         ws5.cell(row=r_idx, column=c_idx, value=value)

syear = input('What year is it? ')
smonth = input('What month is it? ')

wb.save(syear + '_' + smonth + '.DLL.CatalogingStats.xlsx')
